# -*- coding: utf-8 -*-
"""
物资采购需求计划管理系统 - 后端主程序
运行方式: python app.py
启动后浏览器访问 http://127.0.0.1:5000  (同一局域网访问: http://本机IP:5000)
默认超级管理员账号: admin / admin123 (首次启动自动创建, 请登录后立即修改)
"""
import os
import io
import re
import json
import time
import sqlite3
import datetime
import threading

import db as dbm  # PostgreSQL/SQLite 兼容层（根据 DATABASE_URL 自动选择后端）

from flask import (Flask, request, session, jsonify, render_template,
                   send_file, g)

from werkzeug.security import generate_password_hash, check_password_hash

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None
    Workbook = load_workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates_excel")
DB_PATH = os.path.join(DATA_DIR, "system.db")

for d in (DATA_DIR, EXPORT_DIR, TEMPLATE_DIR):
    os.makedirs(d, exist_ok=True)

app = Flask(__name__)
app.secret_key = "wzcg-xqgl-2026-secret-key-change-me"
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20MB 上传限制

# ---------------------------------------------------------------- 数据库
def get_db():
    if "db" not in g:
        g.db = dbm.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        # 仅 SQLite 模式需要 PRAGMA；PostgreSQL 兼容层会自动忽略
        if not dbm.use_postgres():
            g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    db = dbm.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        name TEXT NOT NULL DEFAULT '',
        department TEXT NOT NULL DEFAULT '',
        role TEXT NOT NULL DEFAULT 'employee',   -- super/admin/approver/employee
        receiver_name TEXT NOT NULL DEFAULT '',   -- 收货人(导出自动带出)
        receiver_phone TEXT NOT NULL DEFAULT '',  -- 联系电话(导出自动带出)
        receiver_address TEXT NOT NULL DEFAULT '', -- 收货地址(导出自动带出)
        status TEXT NOT NULL DEFAULT 'active',   -- active/disabled
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL DEFAULT '',          -- 物料编号
        name TEXT NOT NULL DEFAULT '',          -- 物料描述
        spec TEXT NOT NULL DEFAULT '',          -- 规格型号
        unit TEXT NOT NULL DEFAULT '',          -- 计量单位
        price REAL NOT NULL DEFAULT 0,          -- 标准单价
        ecode TEXT NOT NULL DEFAULT '',         -- 电商编码
        supplier_code TEXT NOT NULL DEFAULT '', -- 供应商编码
        supplier TEXT NOT NULL DEFAULT '',      -- 供应商简称
        status TEXT NOT NULL DEFAULT 'active',  -- active/disabled
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS material_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL DEFAULT '',          -- 物料描述(员工填写)
        spec TEXT NOT NULL DEFAULT '',          -- 规格型号(员工填写)
        unit TEXT NOT NULL DEFAULT '',          -- 计量单位(员工填写)
        remark TEXT NOT NULL DEFAULT '',        -- 用途说明(员工填写)
        code TEXT NOT NULL DEFAULT '',          -- 物料编号(管理员补全)
        price REAL NOT NULL DEFAULT 0,          -- 标准单价(管理员补全)
        ecode TEXT NOT NULL DEFAULT '',         -- 电商编码(管理员补全)
        supplier_code TEXT NOT NULL DEFAULT '',
        supplier TEXT NOT NULL DEFAULT '',
        status TEXT NOT NULL DEFAULT 'pending', -- pending/approved/rejected
        created_by INTEGER NOT NULL,
        created_by_name TEXT NOT NULL DEFAULT '',
        created_at TEXT NOT NULL,
        reviewed_by TEXT NOT NULL DEFAULT '',
        review_comment TEXT NOT NULL DEFAULT '',
        reviewed_at TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS demands (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_no TEXT UNIQUE NOT NULL,
        user_id INTEGER NOT NULL,
        department TEXT NOT NULL DEFAULT '',      -- 填报单位
        fill_date TEXT NOT NULL DEFAULT '',       -- 填报日期
        reporter TEXT NOT NULL DEFAULT '',        -- 填表人
        purpose TEXT NOT NULL DEFAULT '',         -- 申请用途(兼容)
        expect_date TEXT NOT NULL DEFAULT '',     -- 需求日期(导出时管理员填)
        remark TEXT NOT NULL DEFAULT '',
        receiver_name TEXT NOT NULL DEFAULT '',    -- 收货人
        receiver_phone TEXT NOT NULL DEFAULT '',   -- 联系电话
        receiver_address TEXT NOT NULL DEFAULT '', -- 收货地址
        status TEXT NOT NULL DEFAULT 'draft',   -- draft/pending/approved/rejected/withdrawn
        created_at TEXT NOT NULL,
        submitted_at TEXT NOT NULL DEFAULT '',
        approved_by TEXT NOT NULL DEFAULT '',
        approved_at TEXT NOT NULL DEFAULT '',
        approve_comment TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS demand_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        demand_id INTEGER NOT NULL,
        seq INTEGER NOT NULL DEFAULT 0,
        material_code TEXT NOT NULL DEFAULT '',
        material_name TEXT NOT NULL DEFAULT '',
        spec TEXT NOT NULL DEFAULT '',
        unit TEXT NOT NULL DEFAULT '',
        price REAL NOT NULL DEFAULT 0,
        quantity REAL NOT NULL DEFAULT 0,
        amount REAL NOT NULL DEFAULT 0,
        ecode TEXT NOT NULL DEFAULT '',
        supplier_code TEXT NOT NULL DEFAULT '',
        supplier TEXT NOT NULL DEFAULT '',
        erp_no TEXT NOT NULL DEFAULT '',          -- ERP需求计划号(导出时填写并回写)
        ec_order_no TEXT NOT NULL DEFAULT ''      -- 电商订单号(导出时填写并回写)
    );
    CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL DEFAULT '',
        action TEXT NOT NULL DEFAULT '',
        detail TEXT NOT NULL DEFAULT '',
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS dept_receivers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department TEXT NOT NULL DEFAULT '',
        receiver_name TEXT NOT NULL DEFAULT '',
        receiver_phone TEXT NOT NULL DEFAULT '',
        receiver_address TEXT NOT NULL DEFAULT '',
        updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS address_book (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        receiver_name TEXT NOT NULL DEFAULT '',
        receiver_phone TEXT NOT NULL DEFAULT '',
        receiver_address TEXT NOT NULL DEFAULT '',
        is_default INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT NOT NULL
    );
    CREATE INDEX IF NOT EXISTS idx_demand_items_demand ON demand_items(demand_id);
    CREATE INDEX IF NOT EXISTS idx_demands_status ON demands(status);
    """)
    # 首次启动创建超级管理员
    cur = db.execute("SELECT COUNT(*) AS c FROM users")
    if cur.fetchone()["c"] == 0:
        db.execute(
            "INSERT INTO users (username,password,name,department,role,status,created_at) VALUES (?,?,?,?,?,?,?)",
            ("admin", generate_password_hash("admin123"), "超级管理员", "物资管理部",
             "super", "active", now_str()),
        )
    # 旧库自动迁移：为 demands / users / demand_items 表补充字段
    for table, cols in (
        ("demands", ("receiver_name", "receiver_phone", "receiver_address", "fill_date", "reporter")),
        ("users", ("receiver_name", "receiver_phone", "receiver_address")),
        ("demand_items", ("erp_no", "ec_order_no")),
    ):
        try:
            exist_cols = [r["name"] for r in db.execute("PRAGMA table_info(%s)" % table).fetchall()]
            for col in cols:
                if col not in exist_cols:
                    db.execute("ALTER TABLE %s ADD COLUMN %s TEXT NOT NULL DEFAULT ''" % (table, col))
        except Exception:
            # PostgreSQL 模式下不会抛 sqlite3.Error，统一吞掉迁移错误（幂等）
            pass
    db.commit()
    db.close()

def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def parse_float(v):
    """安全转 float，失败返回 0。"""
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0

def today_str():
    return datetime.datetime.now().strftime("%Y%m%d")

def add_log(username, action, detail=""):
    db = get_db()
    db.execute("INSERT INTO logs (username,action,detail,created_at) VALUES (?,?,?,?)",
               (username, action, detail, now_str()))
    db.commit()

def next_order_no():
    db = get_db()
    prefix = "XQ" + today_str() + "-"
    row = db.execute(
        "SELECT order_no FROM demands WHERE order_no LIKE ? ORDER BY order_no DESC LIMIT 1",
        (prefix + "%",)).fetchone()
    if row and row["order_no"]:
        n = int(row["order_no"].rsplit("-", 1)[1]) + 1
    else:
        n = 1
    return prefix + "%03d" % n

# ---------------------------------------------------------------- 权限工具
def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "uid" not in session:
            return jsonify({"ok": False, "msg": "未登录或登录已过期"}), 401
        return f(*args, **kwargs)
    return wrapper

def role_required(*roles):
    from functools import wraps
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "uid" not in session:
                return jsonify({"ok": False, "msg": "未登录或登录已过期"}), 401
            db = get_db()
            u = db.execute("SELECT * FROM users WHERE id=? AND status='active'",
                           (session["uid"],)).fetchone()
            if not u:
                session.clear()
                return jsonify({"ok": False, "msg": "账号不存在或已停用"}), 401
            g.user = u
            if u["role"] not in roles:
                return jsonify({"ok": False, "msg": "无权限执行该操作"}), 403
            return f(*args, **kwargs)
        return wrapper
    return deco

def cur_user():
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id=?", (session["uid"],)).fetchone()

def demand_to_dict(d, items=None, with_items=False):
    d = dict(d)
    if with_items:
        d["items"] = [dict(i) for i in items] if items else []
        d["total"] = round(sum(i["amount"] for i in (items or [])), 2)
    return d

# ---------------------------------------------------------------- 页面
@app.route("/")
def index():
    return render_template("login.html")

@app.route("/app")
@login_required
def app_page():
    return render_template("main.html")

# ---------------------------------------------------------------- 认证
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not u or not check_password_hash(u["password"], password):
        return jsonify({"ok": False, "msg": "用户名或密码错误"})
    if u["status"] != "active":
        return jsonify({"ok": False, "msg": "账号已被停用，请联系管理员"})
    session.clear()
    session["uid"] = u["id"]
    session["role"] = u["role"]
    session["name"] = u["name"]
    add_log(u["username"], "登录系统")
    return jsonify({"ok": True, "user": {
        "id": u["id"], "username": u["username"], "name": u["name"],
        "department": u["department"], "role": u["role"]}})

@app.route("/api/logout", methods=["POST"])
@login_required
def api_logout():
    add_log(session.get("name", ""), "退出系统")
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
@login_required
def api_me():
    u = cur_user()
    return jsonify({"ok": True, "user": dict(u)})

@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    data = request.get_json(force=True) or {}
    old_pwd = data.get("old_password") or ""
    new_pwd = data.get("new_password") or ""
    if len(new_pwd) < 6:
        return jsonify({"ok": False, "msg": "新密码长度不能少于6位"})
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (session["uid"],)).fetchone()
    if not check_password_hash(u["password"], old_pwd):
        return jsonify({"ok": False, "msg": "原密码错误"})
    db.execute("UPDATE users SET password=? WHERE id=?",
               (generate_password_hash(new_pwd), session["uid"]))
    db.commit()
    add_log(u["username"], "修改密码")
    return jsonify({"ok": True, "msg": "密码修改成功"})

# ---------------------------------------------------------------- 用户管理(开发者/高级管理员)
@app.route("/api/users", methods=["GET"])
@role_required("super", "admin")
def api_users():
    db = get_db()
    # 高级管理员看不到开发者账号
    if g.user["role"] == "super":
        sql = ("SELECT id,username,name,department,role,receiver_name,receiver_phone,"
               "receiver_address,status,created_at FROM users ORDER BY id")
        rows = db.execute(sql).fetchall()
    else:
        sql = ("SELECT id,username,name,department,role,receiver_name,receiver_phone,"
               "receiver_address,status,created_at FROM users WHERE role<>'super' ORDER BY id")
        rows = db.execute(sql).fetchall()
    return jsonify({"ok": True, "users": [dict(r) for r in rows]})

@app.route("/api/users", methods=["POST"])
@role_required("super", "admin")
def api_user_create():
    data = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or "123456"
    name = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    role = data.get("role") or "employee"
    if role not in ("super", "admin", "approver", "employee"):
        return jsonify({"ok": False, "msg": "角色不合法"})
    # 仅开发者可创建开发者账号
    if role == "super" and g.user["role"] != "super":
        return jsonify({"ok": False, "msg": "仅开发者可创建开发者账号"})
    if not username or not re.match(r"^[\w\u4e00-\u9fa5]{2,30}$", username):
        return jsonify({"ok": False, "msg": "用户名需为2-30位字符(中文/字母/数字/下划线)"})
    if len(password) < 6:
        return jsonify({"ok": False, "msg": "初始密码长度不能少于6位"})
    db = get_db()
    if db.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({"ok": False, "msg": "用户名已存在"})
    db.execute(
        "INSERT INTO users (username,password,name,department,role,status,created_at) VALUES (?,?,?,?,?,?,?)",
        (username, generate_password_hash(password), name, department, role, "active", now_str()))
    db.commit()
    add_log(g.user["username"], "创建账号", username + "/" + name)
    return jsonify({"ok": True, "msg": "账号创建成功"})

@app.route("/api/users/<int:uid>", methods=["PUT"])
@role_required("super", "admin")
def api_user_update(uid):
    """编辑账号：姓名/部门/角色/收货人/联系电话/收货地址（收货信息导出时自动带出）"""
    data = request.get_json(force=True) or {}
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        return jsonify({"ok": False, "msg": "账号不存在"})
    # 开发者账号仅开发者可管理
    if u["role"] == "super" and g.user["role"] != "super":
        return jsonify({"ok": False, "msg": "无权限管理开发者账号"})
    role = data.get("role") or u["role"]
    if role not in ("super", "admin", "approver", "employee"):
        return jsonify({"ok": False, "msg": "角色不合法"})
    # 仅开发者可分配开发者角色（防止越权/自我提权）
    if role == "super" and g.user["role"] != "super":
        return jsonify({"ok": False, "msg": "仅开发者可分配开发者角色"})
    db.execute(
        "UPDATE users SET name=?,department=?,role=?,receiver_name=?,receiver_phone=?,receiver_address=? WHERE id=?",
        ((data.get("name") or "").strip(), (data.get("department") or "").strip(), role,
         (data.get("receiver_name") or "").strip(), (data.get("receiver_phone") or "").strip(),
         (data.get("receiver_address") or "").strip(), uid))
    db.commit()
    add_log(g.user["username"], "编辑账号", u["username"])
    return jsonify({"ok": True, "msg": "账号信息已更新"})


@app.route("/api/users/<int:uid>/reset-password", methods=["POST"])
@role_required("super", "admin")
def api_user_reset(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        return jsonify({"ok": False, "msg": "账号不存在"})
    # 开发者账号仅开发者可管理
    if u["role"] == "super" and g.user["role"] != "super":
        return jsonify({"ok": False, "msg": "无权限管理开发者账号"})
    db.execute("UPDATE users SET password=? WHERE id=?",
               (generate_password_hash("123456"), uid))
    db.commit()
    add_log(g.user["username"], "重置密码", u["username"])
    return jsonify({"ok": True, "msg": "密码已重置为 123456"})

@app.route("/api/users/<int:uid>/toggle", methods=["POST"])
@role_required("super", "admin")
def api_user_toggle(uid):
    if uid == g.user["id"]:
        return jsonify({"ok": False, "msg": "不能停用自己的账号"})
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        return jsonify({"ok": False, "msg": "账号不存在"})
    # 开发者账号仅开发者可管理
    if u["role"] == "super" and g.user["role"] != "super":
        return jsonify({"ok": False, "msg": "无权限管理开发者账号"})
    new_status = "disabled" if u["status"] == "active" else "active"
    db.execute("UPDATE users SET status=? WHERE id=?", (new_status, uid))
    db.commit()
    add_log(g.user["username"], "启停账号", u["username"] + "->" + new_status)
    return jsonify({"ok": True, "msg": "操作成功"})

# ---------------------------------------------------------------- 通用收货地址簿（导出时按顺序匹配）
@app.route("/api/address-book", methods=["GET"])
@role_required("admin", "super")
def api_address_book():
    """返回通用收货地址簿列表"""
    db = get_db()
    rows = db.execute("SELECT id, receiver_name, receiver_phone, receiver_address, is_default, updated_at "
                      "FROM address_book ORDER BY id").fetchall()
    return jsonify({"ok": True, "addresses": [dict(r) for r in rows]})


@app.route("/api/address-book", methods=["POST"])
@role_required("admin", "super")
def api_address_book_save():
    """覆盖保存通用收货地址簿（清空后重新插入，保证顺序与前端一致）"""
    data = request.get_json(force=True) or {}
    addresses = data.get("addresses") or []
    db = get_db()
    db.execute("DELETE FROM address_book")
    saved = 0
    for a in addresses:
        receiver = (a.get("receiver_name") or "").strip()
        phone = (a.get("receiver_phone") or "").strip()
        address = (a.get("receiver_address") or "").strip()
        if not (receiver or phone or address):
            continue
        db.execute("INSERT INTO address_book (receiver_name, receiver_phone, receiver_address, is_default, updated_at) "
                   "VALUES (?, ?, ?, ?, ?)", (receiver, phone, address, 0, now_str()))
        saved += 1
    db.commit()
    add_log(g.user["username"], "维护收货地址簿", "保存 %d 条" % saved)
    return jsonify({"ok": True, "msg": "收货地址已保存（%d 条）" % saved})

# ---------------------------------------------------------------- 部门收货信息配置（导出时自动匹配）
@app.route("/api/dept-receivers", methods=["GET"])
@role_required("admin", "super")
def api_dept_receivers():
    """返回已配置的部门收货信息，以及系统内出现过但尚未配置的部门"""
    db = get_db()
    rows = db.execute("SELECT * FROM dept_receivers ORDER BY department").fetchall()
    configured = {r["department"] for r in rows}
    depts = set()
    for r in db.execute("SELECT DISTINCT department FROM demands WHERE department<>''").fetchall():
        depts.add(r["department"])
    for r in db.execute("SELECT DISTINCT department FROM users WHERE department<>''").fetchall():
        depts.add(r["department"])
    return jsonify({"ok": True, "receivers": [dict(r) for r in rows],
                    "departments": sorted(depts - configured)})


@app.route("/api/dept-receivers", methods=["POST"])
@role_required("admin", "super")
def api_dept_receivers_save():
    """批量保存部门收货信息配置（按部门 upsert，未填写任何字段的行跳过）"""
    data = request.get_json(force=True) or {}
    items = data.get("items") or []
    db = get_db()
    saved = 0
    for it in items:
        department = (it.get("department") or "").strip()
        if not department:
            continue
        receiver = (it.get("receiver_name") or "").strip()
        phone = (it.get("receiver_phone") or "").strip()
        address = (it.get("receiver_address") or "").strip()
        if not (receiver or phone or address):
            continue
        exist = db.execute("SELECT id FROM dept_receivers WHERE department=?",
                           (department,)).fetchone()
        if exist:
            db.execute("UPDATE dept_receivers SET receiver_name=?,receiver_phone=?,receiver_address=?,"
                       "updated_at=? WHERE id=?",
                       (receiver, phone, address, now_str(), exist["id"]))
        else:
            db.execute("INSERT INTO dept_receivers (department,receiver_name,receiver_phone,"
                       "receiver_address,updated_at) VALUES (?,?,?,?,?)",
                       (department, receiver, phone, address, now_str()))
        saved += 1
    db.commit()
    add_log(g.user["username"], "配置部门收货信息", "共保存 %d 条" % saved)
    return jsonify({"ok": True, "msg": "部门收货信息已保存（%d 条）" % saved})

# ---------------------------------------------------------------- 物料库
def search_materials(db, kw, limit=50, sort_by="id", order="desc"):
    """物料查询。sort_by 支持 id/code/name，order 支持 asc/desc。"""
    sql = "SELECT * FROM materials WHERE status='active'"
    args = []
    if kw:
        like = "%" + kw.strip() + "%"
        sql += " AND (name LIKE ? OR code LIKE ? OR spec LIKE ? OR supplier LIKE ?)"
        args = [like, like, like, like]
    order_sql = {"code": "code", "name": "name", "id": "id"}.get(sort_by, "id")
    order_dir = "ASC" if str(order).lower() == "asc" else "DESC"
    sql += " ORDER BY %s %s, id DESC LIMIT ?" % (order_sql, order_dir)
    args.append(limit)
    return db.execute(sql, args).fetchall()

@app.route("/api/materials", methods=["GET"])
@login_required
def api_materials():
    kw = request.args.get("kw", "")
    sort_by = request.args.get("sort_by", "id")
    order = request.args.get("order", "desc")
    db = get_db()
    rows = search_materials(db, kw, limit=100, sort_by=sort_by, order=order)
    return jsonify({"ok": True, "materials": [dict(r) for r in rows]})

@app.route("/api/materials", methods=["POST"])
@role_required("admin", "super")
def api_material_create():
    data = request.get_json(force=True) or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "msg": "物料描述不能为空"})
    db = get_db()
    db.execute(
        "INSERT INTO materials (code,name,spec,unit,price,ecode,supplier_code,supplier,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (code, name, (data.get("spec") or "").strip(), (data.get("unit") or "").strip(),
         float(data.get("price") or 0), (data.get("ecode") or "").strip(),
         (data.get("supplier_code") or "").strip(), (data.get("supplier") or "").strip(),
         "active", now_str()))
    db.commit()
    add_log(g.user["username"], "新增物料", name)
    return jsonify({"ok": True, "msg": "物料已新增"})

@app.route("/api/materials/<int:mid>", methods=["PUT"])
@role_required("admin", "super")
def api_material_update(mid):
    data = request.get_json(force=True) or {}
    db = get_db()
    m = db.execute("SELECT * FROM materials WHERE id=?", (mid,)).fetchone()
    if not m:
        return jsonify({"ok": False, "msg": "物料不存在"})
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "msg": "物料描述不能为空"})
    db.execute(
        "UPDATE materials SET code=?,name=?,spec=?,unit=?,price=?,ecode=?,supplier_code=?,supplier=? WHERE id=?",
        ((data.get("code") or "").strip(), name, (data.get("spec") or "").strip(),
         (data.get("unit") or "").strip(), float(data.get("price") or 0),
         (data.get("ecode") or "").strip(), (data.get("supplier_code") or "").strip(),
         (data.get("supplier") or "").strip(), mid))
    db.commit()
    add_log(g.user["username"], "修改物料", name)
    return jsonify({"ok": True, "msg": "物料已更新"})

@app.route("/api/materials/<int:mid>/disable", methods=["POST"])
@role_required("admin", "super")
def api_material_disable(mid):
    db = get_db()
    m = db.execute("SELECT * FROM materials WHERE id=?", (mid,)).fetchone()
    if not m:
        return jsonify({"ok": False, "msg": "物料不存在"})
    new_status = "disabled" if m["status"] == "active" else "active"
    db.execute("UPDATE materials SET status=? WHERE id=?", (new_status, mid))
    db.commit()
    add_log(g.user["username"], "上下架物料", m["name"])
    return jsonify({"ok": True, "msg": "操作成功"})

@app.route("/api/materials/batch-delete", methods=["POST"])
@role_required("admin", "super")
def api_material_batch_delete():
    """批量删除物料（物理删除）。已下架的物料可删除；若物料在需求单中被引用，
    提示但允许删除（需求单中保存的是文本快照，不受影响）。"""
    data = request.get_json(force=True) or {}
    ids = data.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"ok": False, "msg": "未选择要删除的物料"})
    ids = [int(i) for i in ids if str(i).isdigit()]
    if not ids:
        return jsonify({"ok": False, "msg": "未选择要删除的物料"})
    db = get_db()
    placeholders = ",".join("?" for _ in ids)
    rows = db.execute(
        "SELECT id,name FROM materials WHERE id IN (%s)" % placeholders, ids).fetchall()
    if not rows:
        return jsonify({"ok": False, "msg": "未找到对应物料"})
    # 统计被需求单引用的物料（提示用）
    ref_codes = set()
    for r in rows:
        ref_codes.add(r["name"])
    used = set()
    if ref_codes:
        for nm in ref_codes:
            c = db.execute(
                "SELECT COUNT(*) AS c FROM demand_items WHERE material_name=?", (nm,)).fetchone()["c"]
            if c:
                used.add(nm)
    db.execute("DELETE FROM materials WHERE id IN (%s)" % placeholders, ids)
    db.commit()
    names = "、".join(r["name"] for r in rows[:5])
    if len(rows) > 5:
        names += " 等%d条" % len(rows)
    add_log(g.user["username"], "批量删除物料", "删除%d条: %s" % (len(rows), names))
    msg = "已删除 %d 条物料" % len(rows)
    if used:
        msg += "；其中 %s 曾在需求单中出现过（需求单保留原快照，不受影响）" % "、".join(list(used)[:5])
    return jsonify({"ok": True, "msg": msg, "deleted": len(rows)})

@app.route("/api/materials/import", methods=["POST"])
@role_required("admin", "super")
def api_material_import():
    if load_workbook is None:
        return jsonify({"ok": False, "msg": "服务端缺少 openpyxl，无法导入"})
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "msg": "未选择文件"})
    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"ok": False, "msg": "文件解析失败: " + str(e)})
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"ok": False, "msg": "文件为空"})

    # 表头映射（兼容多种命名，含全角/半角括号变体，如"供应商(缩写)"）
    key_map = {
        "物料编号": "code", "物料编码": "code", "编码": "code", "商品编码": "code", "电商物料编号": "code",
        "物料描述": "name", "物料名称": "name", "名称": "name", "描述": "name", "商品名称": "name",
        "规格型号": "spec", "规格": "spec", "型号": "spec",
        "计量单位": "unit", "单位": "unit", "计量单位名称": "unit",
        "标准单价": "price", "单价": "price", "价格": "price", "标准价格": "price", "含税单价": "price",
        "电商编码": "ecode", "电商编号": "ecode", "商品编码ID": "ecode",
        "供应商编码": "supplier_code", "供应商编号": "supplier_code",
        "供应商": "supplier", "供应商简称": "supplier", "供应商缩写": "supplier",
    }

    def norm_header(h):
        """表头归一化：去空白、全角转半角、去掉括号及括号内内容"""
        h = (h or "").strip()
        out = []
        for ch in h:
            code = ord(ch)
            if code in (0x3000, 0x20):  # 全角/半角空格
                continue
            if 0xFF01 <= code <= 0xFF5E:  # 全角转半角
                out.append(chr(code - 0xFEE0))
            else:
                out.append(ch)
        s = "".join(out)
        s = re.sub(r"[\(（][^)）]*[\)）]", "", s)  # 去掉括号及内容
        return s

    # 在前5行内查找表头行（包含"物料描述/物料名称/商品名称"任一词）
    header_row_idx = 0
    for ri in range(min(5, len(rows))):
        for v in rows[ri]:
            if v and norm_header(str(v)) in ("物料描述", "物料名称", "商品名称", "名称", "描述"):
                header_row_idx = ri
                break
        if header_row_idx:
            break

    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    mapping = {}
    for i, h in enumerate(header):
        f = key_map.get(norm_header(h))
        if f and f not in mapping:
            mapping[f] = i
    if "name" not in mapping:
        return jsonify({"ok": False, "msg": "未识别到『物料描述/物料名称』列，请确认表格首行为表头（如：物料编号、物料描述、规格型号、计量单位、标准单价、电商编码、供应商编码、供应商）"})

    # 非物料数据行特征（合计、审核、签字、备注等模板尾部区域）
    NON_DATA_KEYWORDS = ("合计", "小计", "总计", "审核", "审批", "意见", "编制",
                         "制表", "备注", "说明", "签字", "盖章", "负责人", "经办人", "供应商盖章", "日期")

    db = get_db()

    def getv(row, field):
        ci = mapping.get(field)
        if ci is None or ci >= len(row):
            return ""
        v = row[ci]
        return str(v).strip() if v is not None else ""

    def parse_price(v):
        try:
            return float(v) if v else 0
        except ValueError:
            return 0

    added = skipped = missing = non_data = 0
    dup_skipped = 0    # 编号/描述/单价/供应商完全一致而自动跳过的条数
    problems = []      # 硬问题（缺描述等）
    conflicts = []     # 待用户决策的冲突（编号或描述与库内重复）
    # 记录本次已插入的物料，用于检测文件内部自身的编号/描述重复
    seen_mats = []
    for idx, row in enumerate(rows[header_row_idx + 1:], header_row_idx + 2):
        if not any(v not in (None, "") for v in row):
            continue
        row_text = " ".join(str(v) for v in row if v is not None)
        # 非数据行：含模板固定文字（合计/审核/备注等）
        if any(k in row_text for k in NON_DATA_KEYWORDS):
            non_data += 1
            continue

        name = getv(row, "name")
        code = getv(row, "code")
        if not name:
            # 若整行无任何物料核心字段，视为疑似非数据行（如单独金额行）
            if not (code or getv(row, "ecode") or getv(row, "spec") or getv(row, "unit") or getv(row, "supplier")):
                non_data += 1
                continue
            missing += 1
            skipped += 1
            if len(problems) < 100:
                problems.append({"row": idx, "reason": "缺少物料描述"})
            continue

        new_mat = {
            "code": code, "name": name, "spec": getv(row, "spec"),
            "unit": getv(row, "unit"), "price": parse_price(getv(row, "price")),
            "ecode": getv(row, "ecode"), "supplier_code": getv(row, "supplier_code"),
            "supplier": getv(row, "supplier"),
        }
        # 去重判定：物料编号相同（且编号非空）或 物料描述相同，即视为潜在重复，
        # 是否保留由用户在弹窗中决策（多维度对比编号/描述/单价/电商编码/供应商）
        conds = []
        args = []
        if code:
            conds.append("(code=? AND code!='')")
            args.append(code)
        conds.append("name=?")
        args.append(name)
        exist = db.execute(
            "SELECT id,code,name,spec,unit,price,ecode,supplier_code,supplier,status FROM materials WHERE (%s) ORDER BY id" %
            " OR ".join(conds), args).fetchall()
        # 叠加本次文件内已插入/已确认的物料，识别文件自身重复
        file_dup = [s for s in seen_mats if (code and s["code"] == code) or s["name"] == name]
        existing = [dict(r) for r in exist] + file_dup
        if existing:
            # 完全一致判定：物料编号、物料描述、单价、供应商均相同 → 视为确定的重复物料，
            # 系统自动保留库内已有记录，静默跳过本次导入行，不弹窗
            new_price = new_mat["price"]
            new_supplier = (new_mat["supplier"] or "").strip()
            is_exact_dup = False
            for s in existing:
                if (s["code"] == code and s["name"] == name
                        and parse_price(s["price"]) == new_price
                        and (s.get("supplier") or "").strip() == new_supplier):
                    is_exact_dup = True
                    break
            if is_exact_dup:
                dup_skipped += 1
                skipped += 1
                continue
            conflicts.append({
                "row": idx, "new": new_mat, "existing": existing,
            })
            skipped += 1
            continue
        cur = db.execute(
            "INSERT INTO materials (code,name,spec,unit,price,ecode,supplier_code,supplier,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (new_mat["code"], new_mat["name"], new_mat["spec"], new_mat["unit"],
             new_mat["price"], new_mat["ecode"], new_mat["supplier_code"],
             new_mat["supplier"], "active", now_str()))
        new_id = cur.lastrowid
        if new_id is None:
            r = db.execute("SELECT COALESCE(MAX(id),0) AS i FROM materials").fetchone()
            new_id = r["i"]
        seen_mats.append(dict(new_mat, id=new_id))
        added += 1
    db.commit()
    add_log(g.user["username"], "批量导入物料",
            "新增%d条 完全重复自动跳过%d条 冲突待确认%d条 缺描述%d条 非数据行%d条" % (added, dup_skipped, len(conflicts), missing, non_data))
    if conflicts:
        return jsonify({
            "ok": True, "added": added, "conflicts": conflicts,
            "missing": missing, "non_data": non_data, "problems": problems,
            "msg": "解析完成：直接入库 %d 条，完全重复自动跳过 %d 条，发现 %d 条与物料库重复（请逐条确认），缺描述 %d 条，跳过非物料行 %d 条" %
                   (added, dup_skipped, len(conflicts), missing, non_data),
        })
    return jsonify({
        "ok": True, "added": added, "conflicts": [],
        "missing": missing, "non_data": non_data, "problems": problems,
        "msg": "导入完成：新增 %d 条，完全重复自动跳过 %d 条，缺描述 %d 条，跳过非物料行 %d 条" %
               (added, dup_skipped, missing, non_data),
    })

@app.route("/api/materials/import-confirm", methods=["POST"])
@role_required("admin", "super")
def api_material_import_confirm():
    """导入冲突确认：接收用户的逐条决策。
    每项决策：{"row":行号, "action": "new"|"existing"|"all", "new": {...}, "existing_ids": [...]}
      - new      保留本次导入的物料（新物料入库，库内重复物料下架）
      - existing 保留库内已有（本次导入不入库）
      - all      都保留（新物料入库，库内物料保持不动）
    """
    data = request.get_json(force=True) or {}
    decisions = data.get("decisions") or []
    if not isinstance(decisions, list) or not decisions:
        return jsonify({"ok": False, "msg": "未收到决策数据"})
    db = get_db()
    inserted = kept_new = kept_existing = 0
    for d in decisions:
        action = d.get("action")
        new_mat = d.get("new") or {}
        if action == "new":
            # 新物料入库，库内重复的下架
            db.execute(
                "INSERT INTO materials (code,name,spec,unit,price,ecode,supplier_code,supplier,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (new_mat.get("code", ""), new_mat.get("name", ""), new_mat.get("spec", ""),
                 new_mat.get("unit", ""), parse_float(new_mat.get("price")),
                 new_mat.get("ecode", ""), new_mat.get("supplier_code", ""),
                 new_mat.get("supplier", ""), "active", now_str()))
            for eid in (d.get("existing_ids") or []):
                if str(eid).isdigit():
                    db.execute("UPDATE materials SET status='disabled' WHERE id=?", (int(eid),))
            inserted += 1
            kept_new += 1
        elif action == "all":
            db.execute(
                "INSERT INTO materials (code,name,spec,unit,price,ecode,supplier_code,supplier,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (new_mat.get("code", ""), new_mat.get("name", ""), new_mat.get("spec", ""),
                 new_mat.get("unit", ""), parse_float(new_mat.get("price")),
                 new_mat.get("ecode", ""), new_mat.get("supplier_code", ""),
                 new_mat.get("supplier", ""), "active", now_str()))
            inserted += 1
            kept_new += 1
        else:  # existing
            kept_existing += 1
    db.commit()
    add_log(g.user["username"], "批量导入物料(冲突确认)",
            "新增%d条 保留库内%d条" % (inserted, kept_existing))
    return jsonify({
        "ok": True, "inserted": inserted,
        "kept_existing": kept_existing,
        "msg": "已按您的选择处理完成：新增入库 %d 条，保留库内已有 %d 条" %
               (inserted, kept_existing),
    })

# ---------------------------------------------------------------- 新物料入库申请
@app.route("/api/material-requests", methods=["GET"])
@role_required("admin", "super")
def api_material_requests():
    status = request.args.get("status", "")
    db = get_db()
    sql = "SELECT * FROM material_requests"
    args = []
    if status:
        sql += " WHERE status=?"
        args.append(status)
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, args).fetchall()
    return jsonify({"ok": True, "requests": [dict(r) for r in rows]})

@app.route("/api/material-requests", methods=["POST"])
@login_required
def api_material_request_create():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "msg": "请填写物料描述"})
    u = cur_user()
    db = get_db()
    db.execute(
        "INSERT INTO material_requests (name,spec,unit,remark,status,created_by,created_by_name,created_at) VALUES (?,?,?,?,?,?,?,?)",
        (name, (data.get("spec") or "").strip(), (data.get("unit") or "").strip(),
         (data.get("remark") or "").strip(), "pending", u["id"], u["name"], now_str()))
    db.commit()
    add_log(u["username"], "提交新物料申请", name)
    return jsonify({"ok": True, "msg": "新物料申请已提交，待管理员审核"})

@app.route("/api/material-requests/<int:rid>", methods=["PUT"])
@role_required("admin", "super")
def api_material_request_review(rid):
    data = request.get_json(force=True) or {}
    action = data.get("action") or "approved"
    comment = (data.get("comment") or "").strip()
    db = get_db()
    r = db.execute("SELECT * FROM material_requests WHERE id=?", (rid,)).fetchone()
    if not r or r["status"] != "pending":
        return jsonify({"ok": False, "msg": "该申请不存在或已处理"})
    if action == "approved":
        code = (data.get("code") or "").strip()
        price = float(data.get("price") or 0)
        db.execute(
            "INSERT INTO materials (code,name,spec,unit,price,ecode,supplier_code,supplier,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (code, r["name"], r["spec"], r["unit"], price,
             (data.get("ecode") or "").strip(), (data.get("supplier_code") or "").strip(),
             (data.get("supplier") or "").strip(), "active", now_str()))
        db.execute(
            "UPDATE material_requests SET status='approved',code=?,price=?,ecode=?,supplier_code=?,supplier=?,reviewed_by=?,review_comment=?,reviewed_at=? WHERE id=?",
            (code, price, (data.get("ecode") or "").strip(),
             (data.get("supplier_code") or "").strip(), (data.get("supplier") or "").strip(),
             g.user["name"], comment, now_str(), rid))
        msg = "已通过并纳入物料库"
    else:
        db.execute(
            "UPDATE material_requests SET status='rejected',reviewed_by=?,review_comment=?,reviewed_at=? WHERE id=?",
            (g.user["name"], comment, now_str(), rid))
        msg = "已驳回该申请"
    db.commit()
    add_log(g.user["username"], "审核新物料申请", r["name"] + "->" + msg)
    return jsonify({"ok": True, "msg": msg})

# ---------------------------------------------------------------- 需求单
@app.route("/api/demands", methods=["GET"])
@login_required
def api_demands():
    """员工: 自己的单据; 管理员/超管: 全部; 审批人: 待办+已办"""
    u = cur_user()
    role = u["role"]
    status = request.args.get("status", "")
    kw = (request.args.get("kw") or "").strip()
    db = get_db()

    sql = ("SELECT d.*, u.name AS user_name, u.department AS user_department, "
           "(SELECT SUM(i.amount) FROM demand_items i WHERE i.demand_id=d.id) AS total "
           "FROM demands d LEFT JOIN users u ON d.user_id=u.id WHERE 1=1")
    args = []
    if role == "employee":
        sql += " AND d.user_id=?"
        args.append(u["id"])
    elif role == "approver":
        # 审批人: 待审批(所有已提交未审批) + 已审批(自己审批过的)
        if status == "done":
            sql += " AND d.approved_by=?"
            args.append(u["name"])
        else:
            sql += " AND d.status='pending'"
    if status and role != "approver":
        sql += " AND d.status=?"
        args.append(status)
    if kw:
        like = "%" + kw + "%"
        sql += " AND (d.order_no LIKE ? OR d.purpose LIKE ? OR d.department LIKE ?)"
        args += [like, like, like]
    sql += " ORDER BY d.id DESC LIMIT 200"
    rows = db.execute(sql, args).fetchall()
    return jsonify({"ok": True, "demands": [dict(r) for r in rows]})

@app.route("/api/demands/<int:did>", methods=["GET"])
@login_required
def api_demand_detail(did):
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=?", (did,)).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    u = cur_user()
    if u["role"] == "employee" and d["user_id"] != u["id"]:
        return jsonify({"ok": False, "msg": "无权查看该单据"})
    items = db.execute("SELECT * FROM demand_items WHERE demand_id=? ORDER BY seq", (did,)).fetchall()
    d = dict(d)
    # 收货信息自动匹配：单据已有值 > 填报人账号配置 > 部门收货配置，方便管理员核对
    owner = db.execute("SELECT receiver_name,receiver_phone,receiver_address FROM users WHERE id=?",
                       (d["user_id"],)).fetchone()
    if owner:
        d["receiver_name"] = d["receiver_name"] or owner["receiver_name"] or ""
        d["receiver_phone"] = d["receiver_phone"] or owner["receiver_phone"] or ""
        d["receiver_address"] = d["receiver_address"] or owner["receiver_address"] or ""
    if not (d["receiver_name"] or d["receiver_phone"] or d["receiver_address"]):
        dr = db.execute("SELECT receiver_name,receiver_phone,receiver_address FROM dept_receivers "
                        "WHERE department=? LIMIT 1", (d["department"],)).fetchone()
        if dr:
            d["receiver_name"] = d["receiver_name"] or dr["receiver_name"] or ""
            d["receiver_phone"] = d["receiver_phone"] or dr["receiver_phone"] or ""
            d["receiver_address"] = d["receiver_address"] or dr["receiver_address"] or ""
    return jsonify({"ok": True, "demand": demand_to_dict(d, items, with_items=True)})

@app.route("/api/demands/save-draft", methods=["POST"])
@login_required
def api_demand_save_draft():
    """保存草稿或更新草稿。已提交/已审批的单据不可改。"""
    data = request.get_json(force=True) or {}
    did = data.get("id")
    u = cur_user()
    db = get_db()
    items = data.get("items") or []

    if did:
        d = db.execute("SELECT * FROM demands WHERE id=? AND user_id=?", (did, u["id"])).fetchone()
        if not d:
            return jsonify({"ok": False, "msg": "单据不存在"})
        if d["status"] not in ("draft", "rejected", "withdrawn"):
            return jsonify({"ok": False, "msg": "该单据已提交，无法编辑"})
        order_no = d["order_no"]
        # 保留管理员已回写的 ERP需求计划号/电商订单号，避免重新编辑保存时丢失
        old_items = {}
        for oi in db.execute("SELECT * FROM demand_items WHERE demand_id=?", (did,)).fetchall():
            old_items[(oi["material_code"], oi["material_name"])] = oi
        db.execute("DELETE FROM demand_items WHERE demand_id=?", (did,))
    else:
        old_items = {}
        order_no = next_order_no()

    if not items:
        return jsonify({"ok": False, "msg": "单据中没有物料明细"})
    clean_items = []
    for i, it in enumerate(items):
        if not (it.get("material_name") or "").strip():
            continue
        qty = float(it.get("quantity") or 0)
        price = float(it.get("price") or 0)
        mcode = it.get("material_code") or ""
        mname = (it.get("material_name") or "").strip()
        oi = old_items.get((mcode, mname))
        clean_items.append((
            mcode, mname,
            it.get("spec") or "", it.get("unit") or "", price, qty,
            round(price * qty, 2), it.get("ecode") or "",
            it.get("supplier_code") or "", it.get("supplier") or "",
            (oi["erp_no"] if oi else "") or "", (oi["ec_order_no"] if oi else "") or ""))
    if not clean_items:
        return jsonify({"ok": False, "msg": "单据中没有有效的物料明细"})

    department = (data.get("department") or "").strip() or u["department"]
    fill_date = (data.get("fill_date") or "").strip()
    reporter = (data.get("reporter") or "").strip()
    if did:
        db.execute(
            "UPDATE demands SET department=?,fill_date=?,reporter=?,purpose=?,expect_date=?,remark=?,receiver_name=?,receiver_phone=?,receiver_address=? WHERE id=?",
            (department, fill_date, reporter,
             (data.get("purpose") or "").strip(), (data.get("expect_date") or "").strip(),
             (data.get("remark") or "").strip(),
             (data.get("receiver_name") or "").strip(),
             (data.get("receiver_phone") or "").strip(),
             (data.get("receiver_address") or "").strip(), did))
        demand_id = did
    else:
        cur = db.execute(
            "INSERT INTO demands (order_no,user_id,department,fill_date,reporter,purpose,expect_date,remark,receiver_name,receiver_phone,receiver_address,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (order_no, u["id"], department, fill_date, reporter,
             (data.get("purpose") or "").strip(), (data.get("expect_date") or "").strip(),
             (data.get("remark") or "").strip(),
             (data.get("receiver_name") or "").strip(),
             (data.get("receiver_phone") or "").strip(),
             (data.get("receiver_address") or "").strip(),
             "draft", now_str()))
        demand_id = cur.lastrowid

    for seq, it in enumerate(clean_items, 1):
        db.execute(
            "INSERT INTO demand_items (demand_id,seq,material_code,material_name,spec,unit,price,quantity,amount,ecode,supplier_code,supplier,erp_no,ec_order_no) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (demand_id, seq) + it)
    db.commit()
    add_log(u["username"], "保存草稿", order_no)
    return jsonify({"ok": True, "msg": "草稿已保存", "id": demand_id})

@app.route("/api/demands/<int:did>/submit", methods=["POST"])
@login_required
def api_demand_submit(did):
    u = cur_user()
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=? AND user_id=?", (did, u["id"])).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    if d["status"] != "draft":
        return jsonify({"ok": False, "msg": "仅草稿状态可提交"})
    cnt = db.execute("SELECT COUNT(*) AS c FROM demand_items WHERE demand_id=?", (did,)).fetchone()["c"]
    if cnt == 0:
        return jsonify({"ok": False, "msg": "单据中没有物料明细，无法提交"})
    db.execute("UPDATE demands SET status='pending',submitted_at=? WHERE id=?",
               (now_str(), did))
    db.commit()
    add_log(u["username"], "提交需求单", d["order_no"])
    return jsonify({"ok": True, "msg": "已提交，等待审批人处理"})

@app.route("/api/demands/<int:did>/withdraw", methods=["POST"])
@login_required
def api_demand_withdraw(did):
    u = cur_user()
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=? AND user_id=?", (did, u["id"])).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    if d["status"] != "pending":
        return jsonify({"ok": False, "msg": "仅待审批单据可撤回"})
    db.execute("UPDATE demands SET status='withdrawn' WHERE id=?", (did,))
    db.commit()
    add_log(u["username"], "撤回需求单", d["order_no"])
    return jsonify({"ok": True, "msg": "已撤回，可重新编辑提交"})

@app.route("/api/demands/<int:did>/copy", methods=["POST"])
@login_required
def api_demand_copy(did):
    u = cur_user()
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=? AND user_id=?", (did, u["id"])).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    items = db.execute("SELECT * FROM demand_items WHERE demand_id=? ORDER BY seq", (did,)).fetchall()
    if not items:
        return jsonify({"ok": False, "msg": "原单据无物料明细"})
    order_no = next_order_no()
    cur = db.execute(
        "INSERT INTO demands (order_no,user_id,department,fill_date,reporter,purpose,expect_date,remark,receiver_name,receiver_phone,receiver_address,status,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (order_no, u["id"], d["department"], d["fill_date"], d["reporter"],
         d["purpose"], d["expect_date"], d["remark"],
         d["receiver_name"], d["receiver_phone"], d["receiver_address"], "draft", now_str()))
    new_id = cur.lastrowid
    for seq, it in enumerate(items, 1):
        db.execute(
            "INSERT INTO demand_items (demand_id,seq,material_code,material_name,spec,unit,price,quantity,amount,ecode,supplier_code,supplier,erp_no,ec_order_no) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (new_id, seq, it["material_code"], it["material_name"], it["spec"], it["unit"],
             it["price"], it["quantity"], it["amount"], it["ecode"], it["supplier_code"], it["supplier"],
             it["erp_no"] or "", it["ec_order_no"] or ""))
    db.commit()
    add_log(u["username"], "复制新建需求单", d["order_no"] + "->" + order_no)
    return jsonify({"ok": True, "msg": "已复制为草稿，可修改后提交", "id": new_id})

@app.route("/api/demands/<int:did>", methods=["DELETE"])
@login_required
def api_demand_delete(did):
    u = cur_user()
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=? AND user_id=?", (did, u["id"])).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    if d["status"] != "draft":
        return jsonify({"ok": False, "msg": "仅草稿可删除"})
    db.execute("DELETE FROM demand_items WHERE demand_id=?", (did,))
    db.execute("DELETE FROM demands WHERE id=?", (did,))
    db.commit()
    add_log(u["username"], "删除草稿", d["order_no"])
    return jsonify({"ok": True, "msg": "已删除"})

# ---------------------------------------------------------------- 审批
@app.route("/api/demands/<int:did>/approve", methods=["POST"])
@role_required("approver", "admin", "super")
def api_demand_approve(did):
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=?", (did,)).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    if d["status"] != "pending":
        return jsonify({"ok": False, "msg": "该单据不在待审批状态"})
    db.execute("UPDATE demands SET status='approved',approved_by=?,approved_at=?,approve_comment=? WHERE id=?",
               (g.user["name"], now_str(), "", did))
    db.commit()
    add_log(g.user["username"], "审批通过", d["order_no"])
    return jsonify({"ok": True, "msg": "已审批通过"})

@app.route("/api/demands/<int:did>/reject", methods=["POST"])
@role_required("approver", "admin", "super")
def api_demand_reject(did):
    data = request.get_json(force=True) or {}
    comment = (data.get("comment") or "").strip()
    if not comment:
        return jsonify({"ok": False, "msg": "退回必须填写原因"})
    db = get_db()
    d = db.execute("SELECT * FROM demands WHERE id=?", (did,)).fetchone()
    if not d:
        return jsonify({"ok": False, "msg": "单据不存在"})
    if d["status"] != "pending":
        return jsonify({"ok": False, "msg": "该单据不在待审批状态"})
    db.execute("UPDATE demands SET status='rejected',approved_by=?,approved_at=?,approve_comment=? WHERE id=?",
               (g.user["name"], now_str(), comment, did))
    db.commit()
    add_log(g.user["username"], "退回需求单", d["order_no"] + " 原因:" + comment)
    return jsonify({"ok": True, "msg": "已退回，原因已通知填报人"})

# ---------------------------------------------------------------- 汇总导出
# 标准物资需求计划单模板规格（17列，与单位现行模板一致）
STANDARD_HEADERS = ["序号", "物料编号", "物料描述", "单位", "单价", "数量", "计划金额",
                    "需求日期", "提报人", "收货地址", "收货人", "联系电话",
                    "电商编码", "供应商编码", "供应商(缩写)", "ERP需求计划号", "电商订单号"]
STANDARD_WIDTHS = [7.38, 17.13, 45.75, 7.75, 28.5, 13, 13, 13, 13, 13, 13,
                   13, 19, 11.25, 14.63, 22.19, 14.63]
# 模板尾部固定文字
NOTE_LINE1 = "1、物料描述必须为电商平台商品名称，计量单位必须为电商平台售卖计量单位，计划金额必须为电商平台最高价格。"
NOTE_LINE2 = "2、审核意见根据各用户单位情况设置。"


def norm_header(h):
    """表头归一化：去空白、全角转半角、去掉括号及括号内内容"""
    h = (h or "").strip()
    out = []
    for ch in h:
        code = ord(ch)
        if code in (0x3000, 0x20):  # 全角/半角空格
            continue
        if 0xFF01 <= code <= 0xFF5E:  # 全角转半角
            out.append(chr(code - 0xFEE0))
        else:
            out.append(ch)
    s = "".join(out)
    s = re.sub(r"[\(（][^)）]*[\)）]", "", s)  # 去掉括号及内容
    return s


def fmt_date(s):
    """需求日期转模板格式：2026-07-31 -> 2026.7.31"""
    s = (s or "").strip()
    if not s:
        return ""
    m = re.match(r"(\d{4})[-/.年]?(\d{1,2})[-/.月]?(\d{1,2})", s)
    if m:
        return "%s.%s.%s" % (m.group(1), str(int(m.group(2))), str(int(m.group(3))))
    return s


# 导出模板列 → 数据字段映射（物料行 + 单据级字段）
EXPORT_KEY_MAP = {
    "物料编号": "code", "物料编码": "code", "编码": "code",
    "物料描述": "name", "物料名称": "name", "名称": "name", "描述": "name",
    "规格型号": "spec", "规格": "spec", "型号": "spec",
    "计量单位": "unit", "单位": "unit",
    "数量": "quantity", "采购数量": "quantity",
    "标准单价": "price", "单价": "price",
    "金额": "amount", "小计": "amount", "合计金额": "amount", "计划金额": "amount",
    "需求日期": "expect_date", "需求时间": "expect_date",
    "提报人": "user_name", "填报人": "user_name", "申报人": "user_name",
    "收货地址": "receiver_address", "地址": "receiver_address",
    "收货人": "receiver_name", "联系人": "receiver_name",
    "联系电话": "receiver_phone", "电话": "receiver_phone", "手机": "receiver_phone",
    "电商编码": "ecode", "电商编号": "ecode",
    "供应商编码": "supplier_code", "供应商编号": "supplier_code",
    "供应商": "supplier", "供应商简称": "supplier", "供应商缩写": "supplier",
    "ERP需求计划号": "erp_no", "需求计划号": "erp_no",
    "电商订单号": "ec_order_no", "订单号": "ec_order_no",
    "序号": "seq",
}


def build_row(d, it, seq):
    """按字段生成一行数据：it 为物料字典，d 为单据字典"""
    row = {
        "code": it.get("material_code", ""), "name": it.get("material_name", ""),
        "spec": it.get("spec", ""), "unit": it.get("unit", ""),
        "quantity": it.get("quantity", 0), "price": it.get("price", 0),
        "amount": it.get("amount", 0), "ecode": it.get("ecode", ""),
        "supplier_code": it.get("supplier_code", ""), "supplier": it.get("supplier", ""),
        "expect_date": fmt_date(d.get("expect_date", "")),
        "user_name": d.get("user_name", ""),
        "receiver_address": d.get("receiver_address", ""),
        "receiver_name": d.get("receiver_name", ""),
        "receiver_phone": d.get("receiver_phone", ""),
        "erp_no": it.get("erp_no", ""), "ec_order_no": it.get("ec_order_no", ""),
        "seq": seq,
    }
    return row


def get_active_template():
    """读取当前激活的模板文件名"""
    p = os.path.join(TEMPLATE_DIR, ".active")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                name = f.read().strip().lstrip("\ufeff")  # 兼容带BOM的UTF8
            if name and os.path.isfile(os.path.join(TEMPLATE_DIR, name)):
                return name
        except OSError:
            pass
    return None


def find_template():
    """查找当前使用的导出模板文件（优先激活模板，其次目录内任意模板）"""
    name = get_active_template()
    if name:
        return os.path.join(TEMPLATE_DIR, name)
    if os.path.isdir(TEMPLATE_DIR):
        for fn in sorted(os.listdir(TEMPLATE_DIR)):
            if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("."):
                return os.path.join(TEMPLATE_DIR, fn)
    return None
    for f in sorted(os.listdir(TEMPLATE_DIR)):
        if f.lower().endswith((".xlsx", ".xlsm")):
            return os.path.join(TEMPLATE_DIR, f)
    return None

def detect_header_row(ws, key_map):
    """在模板前20行内定位表头行，返回 (行号, 列映射) 或 None"""
    for r in range(1, min(ws.max_row, 20) + 1):
        mapping = {}
        matched = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            key = norm_header(str(v).strip().replace("\n", ""))
            if key in key_map:
                mapping[key_map[key]] = c
                matched += 1
        if matched >= 3:
            return r, mapping
    return None, None


def _thin_border():
    from openpyxl.styles import Border, Side
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        col = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col].width = w


def export_default(demands_map, total_amount, path):
    """无模板时的默认导出：按标准 17 列模板格式生成"""
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "物资需求计划单"
    # 表头
    header_font = Font(name="等线", size=10)
    for c, h in enumerate(STANDARD_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 21.95
    # 数据行
    data_font = Font(name="等线 Light", size=11)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    seq = 0
    r = 2
    for d in demands_map:
        for it in d["items"]:
            seq += 1
            row = build_row(d, it, seq)
            for col_idx, field in enumerate(
                    ["seq", "code", "name", "unit", "price", "quantity", "amount",
                     "expect_date", "user_name", "receiver_address", "receiver_name",
                     "receiver_phone", "ecode", "supplier_code", "supplier",
                     "erp_no", "ec_order_no"], 1):
                cell = ws.cell(row=r, column=col_idx, value=row[field])
                cell.font = data_font
                cell.alignment = data_align
                cell.border = _thin_border()
            ws.row_dimensions[r].height = 40.5
            r += 1
    _append_tail(ws, r, total_amount, data_font, data_align)
    _set_widths(ws, STANDARD_WIDTHS)
    wb.save(path)


def _append_tail(ws, data_end_row, total_amount, font, align):
    """按模板格式追加 合计/审核意见/编制人/备注 固定区域"""
    from openpyxl.styles import Font, Alignment
    r = data_end_row
    # 合计行
    total_cell = ws.cell(row=r, column=7, value=total_amount)  # 计划金额列
    total_cell.font = Font(name="等线 Light", size=11, bold=True)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.border = _thin_border()
    label_cell = ws.cell(row=r, column=1, value="合计")
    label_cell.font = Font(name="等线 Light", size=11, bold=True)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = _thin_border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 40.5
    # 审核意见区域（占4行，A列标签 + D列情况说明）
    r2 = r + 1
    a = ws.cell(row=r2, column=1, value="审核意见")
    a.font = font; a.alignment = align; a.border = _thin_border()
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2 + 3, end_column=2)
    d = ws.cell(row=r2, column=4, value="情况说明")
    d.font = font; d.alignment = align; d.border = _thin_border()
    ws.merge_cells(start_row=r2, start_column=4, end_row=r2 + 3, end_column=12)
    for rr in range(r2, r2 + 4):
        ws.row_dimensions[rr].height = 40.5
    # 编制人行
    r3 = r2 + 4
    e = ws.cell(row=r3, column=1, value="编制人：")
    e.font = font; e.alignment = align; e.border = _thin_border()
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=2)
    ws.row_dimensions[r3].height = 40.5
    # 备注两行
    for offset, txt in ((1, NOTE_LINE1), (2, NOTE_LINE2)):
        rr = r3 + offset
        n = ws.cell(row=rr, column=1, value=txt)
        n.font = font
        n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n.border = _thin_border()
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
        ws.row_dimensions[rr].height = 40.5


def export_with_template(template_path, demands_map, total_amount, path):
    """按标准模板导出：定位表头行，清空旧数据后按模板格式重建数据区与尾部固定区"""
    from openpyxl.styles import Font, Alignment
    wb = load_workbook(template_path)
    ws = wb.active
    header_row, mapping = detect_header_row(ws, EXPORT_KEY_MAP)
    if not header_row:
        return export_default(demands_map, total_amount, path)

    ncols = ws.max_column
    # 1. 清空表头行以下所有内容与合并单元格，避免旧数据残留
    if ws.merged_cells.ranges:
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.value = None
    # 2. 列宽：模板缺失的列用标准宽度兜底
    for i in range(1, ncols + 1):
        col = openpyxl.utils.get_column_letter(i)
        if ws.column_dimensions[col].width is None:
            ws.column_dimensions[col].width = STANDARD_WIDTHS[i - 1] if i <= len(STANDARD_WIDTHS) else 13
    # 3. 写入数据行（按模板映射的列位置）
    data_font = Font(name="等线 Light", size=11)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    seq = 0
    r = header_row + 1
    all_rows = []
    for d in demands_map:
        for it in d["items"]:
            seq += 1
            all_rows.append(build_row(d, it, seq))
    for row in all_rows:
        for field, col in mapping.items():
            val = row.get(field, "")
            if field in ("price", "amount", "quantity") and val not in ("", None):
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = _thin_border()
        ws.row_dimensions[r].height = 40.5
        r += 1
    # 4. 尾部固定区（合计/审核意见/编制人/备注），金额列用模板映射
    amount_col = mapping.get("amount", 7)
    total_cell = ws.cell(row=r, column=amount_col, value=total_amount)
    total_cell.font = Font(name="等线 Light", size=11, bold=True)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.border = _thin_border()
    label_cell = ws.cell(row=r, column=1, value="合计")
    label_cell.font = Font(name="等线 Light", size=11, bold=True)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = _thin_border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 40.5
    r2 = r + 1
    a = ws.cell(row=r2, column=1, value="审核意见")
    a.font = data_font; a.alignment = data_align; a.border = _thin_border()
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2 + 3, end_column=2)
    d = ws.cell(row=r2, column=4, value="情况说明")
    d.font = data_font; d.alignment = data_align; d.border = _thin_border()
    ws.merge_cells(start_row=r2, start_column=4, end_row=r2 + 3, end_column=12)
    for rr in range(r2, r2 + 4):
        ws.row_dimensions[rr].height = 40.5
    r3 = r2 + 4
    e = ws.cell(row=r3, column=1, value="编制人：")
    e.font = data_font; e.alignment = data_align; e.border = _thin_border()
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=2)
    ws.row_dimensions[r3].height = 40.5
    for offset, txt in ((1, NOTE_LINE1), (2, NOTE_LINE2)):
        rr = r3 + offset
        n = ws.cell(row=rr, column=1, value=txt)
        n.font = data_font
        n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n.border = _thin_border()
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
        ws.row_dimensions[rr].height = 40.5
    wb.save(path)

@app.route("/api/demands/export-preview", methods=["POST"])
@role_required("admin", "super")
def api_demands_export_preview():
    """按选中单据/筛选条件生成可编辑的导出行数据（17列标准字段）"""
    data = request.get_json(force=True) or {}
    ids = data.get("ids") or []
    status = data.get("status", "")
    start_date = data.get("start_date", "")
    end_date = data.get("end_date", "")
    department = (data.get("department") or "").strip()

    db = get_db()
    sql = ("SELECT d.*, u.name AS user_name, "
           "u.receiver_name AS u_receiver_name, u.receiver_phone AS u_receiver_phone, "
           "u.receiver_address AS u_receiver_address FROM demands d "
           "LEFT JOIN users u ON d.user_id=u.id WHERE 1=1")
    args = []
    if ids:
        placeholders = ",".join("?" * len(ids))
        sql += " AND d.id IN (" + placeholders + ")"
        args += ids
    else:
        if status:
            sql += " AND d.status=?"
            args.append(status)
        if start_date:
            sql += " AND d.submitted_at >= ?"
            args.append(start_date + " 00:00:00")
        if end_date:
            sql += " AND d.submitted_at <= ?"
            args.append(end_date + " 23:59:59")
        if department:
            sql += " AND d.department=?"
            args.append(department)
    sql += " ORDER BY d.id"
    demands = db.execute(sql, args).fetchall()
    if not demands:
        return jsonify({"ok": False, "msg": "没有符合条件的数据可导出"})

    # 通用收货地址簿：按顺序匹配到预览行
    addr_book = [dict(r) for r in db.execute(
        "SELECT receiver_name, receiver_phone, receiver_address "
        "FROM address_book ORDER BY id").fetchall()]
    # 部门收货信息配置：导出时自动匹配兜底（单据已有值 > 填报人账号配置 > 部门配置）
    dept_recv = {r["department"]: dict(r) for r in db.execute(
        "SELECT department, receiver_name, receiver_phone, receiver_address "
        "FROM dept_receivers WHERE department<>''").fetchall()}

    out = []
    total = 0.0
    for d in demands:
        items = db.execute("SELECT * FROM demand_items WHERE demand_id=? ORDER BY seq",
                           (d["id"],)).fetchall()
        if not items:
            continue
        dr = dept_recv.get(d["department"] or "") or {}
        for it in items:
            it = dict(it)
            amount = round(it["price"] * it["quantity"], 2)
            total += amount
            # 收货信息自动匹配优先级：通用地址簿 > 单据已有记录 > 填报人账号配置 > 部门收货配置
            idx = len(out)
            if idx < len(addr_book) and (addr_book[idx]["receiver_name"] or addr_book[idx]["receiver_phone"] or addr_book[idx]["receiver_address"]):
                address = (addr_book[idx]["receiver_address"] or "").strip()
                receiver = (addr_book[idx]["receiver_name"] or "").strip()
                phone = (addr_book[idx]["receiver_phone"] or "").strip()
            else:
                address = ((d["receiver_address"] or "").strip()
                           or (d["u_receiver_address"] or "").strip()
                           or (dr.get("receiver_address") or "").strip())
                receiver = ((d["receiver_name"] or "").strip()
                            or (d["u_receiver_name"] or "").strip()
                            or (dr.get("receiver_name") or "").strip())
                phone = ((d["receiver_phone"] or "").strip()
                         or (d["u_receiver_phone"] or "").strip()
                         or (dr.get("receiver_phone") or "").strip())
            out.append({
                "seq": len(out) + 1,
                "code": it["material_code"], "name": it["material_name"],
                "unit": it["unit"], "price": it["price"], "quantity": it["quantity"],
                "amount": amount,
                "expect_date": d["expect_date"] or "",
                "reporter": d["reporter"] or d["user_name"] or "",
                "address": address, "receiver": receiver, "phone": phone,
                "ecode": it["ecode"], "supplier_code": it["supplier_code"],
                "supplier": it["supplier"],
                "erp_no": it["erp_no"] or "", "order_no": it["ec_order_no"] or "",
                "demand_order": d["order_no"], "dept": d["department"],
                "demand_id": d["id"], "item_id": it["id"],
            })
    if not out:
        return jsonify({"ok": False, "msg": "所选单据中没有物料明细"})
    return jsonify({"ok": True, "rows": out, "total": round(total, 2),
                    "msg": "共 %d 行物料，请核对后可编辑后导出" % len(out)})


@app.route("/api/demands/export", methods=["POST"])
@role_required("admin", "super")
def api_demands_export():
    """接收管理员编辑后的行数据，按标准模板生成 Excel"""
    data = request.get_json(force=True) or {}
    rows = data.get("rows") or []
    if not rows:
        return jsonify({"ok": False, "msg": "没有可导出的数据行"})
    db = get_db()

    # 数据前置校验（金额重新计算，不信任前端）
    problems = []
    for i, r in enumerate(rows, 1):
        if not (r.get("name") or "").strip():
            problems.append("第%d行缺少物料描述" % i)
        try:
            qty = float(r.get("quantity") or 0)
            price = float(r.get("price") or 0)
        except (TypeError, ValueError):
            problems.append("第%d行数量/单价格式不正确" % i)
            continue
        if qty <= 0 or price < 0:
            problems.append("第%d行数量或单价异常" % i)
    if problems:
        return jsonify({"ok": False, "msg": "以下数据需修正后导出：\n- " + "\n- ".join(problems[:10])})

    demands_map = []
    total = 0.0
    for r in rows:
        qty = float(r.get("quantity") or 0)
        price = float(r.get("price") or 0)
        amount = round(price * qty, 2)
        total += amount
        demands_map.append({
            "order_no": r.get("demand_order", ""), "department": r.get("dept", ""),
            "user_name": r.get("reporter", ""), "expect_date": r.get("expect_date", ""),
            "receiver_name": r.get("receiver", ""), "receiver_phone": r.get("phone", ""),
            "receiver_address": r.get("address", ""),
            "items": [{
                "material_code": r.get("code", ""), "material_name": r.get("name", ""),
                "unit": r.get("unit", ""), "price": price, "quantity": qty,
                "amount": amount, "ecode": r.get("ecode", ""),
                "supplier_code": r.get("supplier_code", ""), "supplier": r.get("supplier", ""),
                "erp_no": r.get("erp_no", ""), "ec_order_no": r.get("order_no", ""),
            }],
        })
    total = round(total, 2)

    template_path = find_template()
    filename = "物资需求计划单_%s.xlsx" % datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    filepath = os.path.join(EXPORT_DIR, filename)
    if template_path:
        try:
            export_with_template(template_path, demands_map, total, filepath)
        except Exception:
            export_default(demands_map, total, filepath)
    else:
        export_default(demands_map, total, filepath)

    # 回写数据库：保存管理员编辑后的 ERP需求计划号/电商订单号/需求日期/收货信息
    try:
        for r in rows:
            iid = r.get("item_id")
            did = r.get("demand_id")
            if iid:
                db.execute("UPDATE demand_items SET erp_no=?,ec_order_no=? WHERE id=?",
                           ((r.get("erp_no") or "").strip(), (r.get("order_no") or "").strip(), iid))
            if did:
                if (r.get("expect_date") or "").strip():
                    db.execute("UPDATE demands SET expect_date=? WHERE id=?",
                               ((r.get("expect_date") or "").strip(), did))
                if (r.get("receiver") or "").strip() or (r.get("phone") or "").strip() or (r.get("address") or "").strip():
                    db.execute("UPDATE demands SET receiver_name=?,receiver_phone=?,receiver_address=? WHERE id=?",
                               ((r.get("receiver") or "").strip(), (r.get("phone") or "").strip(),
                                (r.get("address") or "").strip(), did))
        db.commit()
    except sqlite3.Error:
        pass

    add_log(g.user["username"], "导出需求计划单", filename)
    return jsonify({"ok": True, "msg": "导出成功，共 %d 行物料，合计 %.2f 元" % (len(rows), total),
                    "file": filename})

@app.route("/api/export/<filename>")
@login_required
def api_export_download(filename):
    if not re.match(r"^[\w\-]+\.xlsx?$", filename):
        return jsonify({"ok": False, "msg": "文件名非法"}), 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"ok": False, "msg": "文件不存在或已过期"}), 404
    return send_file(path, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------------------------------------------- 导出模板管理
@app.route("/api/templates", methods=["GET"])
@role_required("admin", "super")
def api_templates():
    files = []
    if os.path.isdir(TEMPLATE_DIR):
        for fn in sorted(os.listdir(TEMPLATE_DIR)):
            if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("."):
                fp = os.path.join(TEMPLATE_DIR, fn)
                files.append({
                    "name": fn, "size": os.path.getsize(fp),
                    "mtime": datetime.datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M"),
                    "active": fn == get_active_template(),
                })
    return jsonify({"ok": True, "templates": files, "active": get_active_template() or ""})


@app.route("/api/templates/upload", methods=["POST"])
@role_required("admin", "super")
def api_templates_upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "msg": "请选择模板文件"})
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "msg": "仅支持 .xlsx / .xls 模板文件"})
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    name = os.path.basename(f.filename)
    dest = os.path.join(TEMPLATE_DIR, name)
    if os.path.exists(dest):
        return jsonify({"ok": False, "msg": "已存在同名模板：" + name + "（请先删除或改名）"})
    f.save(dest)
    add_log(g.user["username"], "上传导出模板", name)
    return jsonify({"ok": True, "msg": "模板已上传：" + name})


@app.route("/api/templates/activate", methods=["POST"])
@role_required("admin", "super")
def api_templates_activate():
    name = (request.get_json(force=True) or {}).get("name", "")
    if not name or not os.path.isfile(os.path.join(TEMPLATE_DIR, name)):
        return jsonify({"ok": False, "msg": "模板不存在"})
    with open(os.path.join(TEMPLATE_DIR, ".active"), "w", encoding="utf-8") as f:
        f.write(name)
    add_log(g.user["username"], "切换导出模板", name)
    return jsonify({"ok": True, "msg": "已切换导出模板：" + name})


@app.route("/api/templates/delete", methods=["POST"])
@role_required("admin", "super")
def api_templates_delete():
    name = (request.get_json(force=True) or {}).get("name", "")
    if not name:
        return jsonify({"ok": False, "msg": "缺少模板名"})
    if name == get_active_template():
        return jsonify({"ok": False, "msg": "当前使用中的模板不能删除，请先切换其他模板"})
    fp = os.path.join(TEMPLATE_DIR, name)
    if os.path.isfile(fp):
        os.remove(fp)
        add_log(g.user["username"], "删除导出模板", name)
    return jsonify({"ok": True, "msg": "模板已删除"})

# ---------------------------------------------------------------- 日志
@app.route("/api/logs", methods=["GET"])
@role_required("super")
def api_logs():
    db = get_db()
    rows = db.execute("SELECT * FROM logs ORDER BY id DESC LIMIT 300").fetchall()
    return jsonify({"ok": True, "logs": [dict(r) for r in rows]})

# ---------------------------------------------------------------- 统计
@app.route("/api/stats")
@login_required
def api_stats():
    u = cur_user()
    db = get_db()
    role = u["role"]
    res = {}
    if role == "employee":
        res["drafts"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE user_id=? AND status='draft'", (u["id"],)).fetchone()["c"]
        res["pending"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE user_id=? AND status='pending'", (u["id"],)).fetchone()["c"]
        res["approved"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE user_id=? AND status='approved'", (u["id"],)).fetchone()["c"]
        res["rejected"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE user_id=? AND status='rejected'", (u["id"],)).fetchone()["c"]
    elif role == "approver":
        res["pending"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE status='pending'").fetchone()["c"]
        res["done"] = db.execute(
            "SELECT COUNT(*) AS c FROM demands WHERE approved_by=?", (u["name"],)).fetchone()["c"]
    else:
        # 高级管理员工作台的在职账号统计不包含开发者(super)账号
        if role == "admin":
            res["total_users"] = db.execute(
                "SELECT COUNT(*) AS c FROM users WHERE status='active' AND role<>'super'").fetchone()["c"]
        else:
            res["total_users"] = db.execute("SELECT COUNT(*) AS c FROM users WHERE status='active'").fetchone()["c"]
        res["materials"] = db.execute("SELECT COUNT(*) AS c FROM materials WHERE status='active'").fetchone()["c"]
        res["pending_demands"] = db.execute("SELECT COUNT(*) AS c FROM demands WHERE status='pending'").fetchone()["c"]
        res["approved_demands"] = db.execute("SELECT COUNT(*) AS c FROM demands WHERE status='approved'").fetchone()["c"]
        res["pending_material_reqs"] = db.execute(
            "SELECT COUNT(*) AS c FROM material_requests WHERE status='pending'").fetchone()["c"]
    return jsonify({"ok": True, "stats": res})

# ---------------------------------------------------------------- 启动
# 首次加载即初始化数据库（云平台用 gunicorn 导入本模块时，__main__ 块不会执行，
# 因此必须在模块层调用 init_db()，确保数据库与表结构、管理员账号始终就绪）
init_db()


if __name__ == "__main__":
    print("=" * 60)
    print("  物资采购需求计划管理系统")
    print("  本机访问:  http://127.0.0.1:5000")
    print("  局域网访问: http://<本机IP>:5000  (同一网络下基层员工用手机/电脑访问)")
    print("  默认账号:  admin / admin123  (请登录后立即修改密码)")
    print("=" * 60)
    # 端口优先读取环境变量 PORT（Render 等云平台会自动注入），默认 5000
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
